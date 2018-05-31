# coding=utf-8
from openpyxl.workbook import Workbook
from openpyxl.styles import Font,colors
import openpyxl,re,string
from PIL import Image

def ten_sixteen(rang):
    tar = hex(rang)
    return tar

def Joining(R,G,B):
    if R > 15:
        R = ten_sixteen(R)[2:4]
    else:
        r = ten_sixteen(R)[2:3]
        R = '0' + r
        print R
    if G > 15:
        G = ten_sixteen(G)[2:4]
    else:
        g = ten_sixteen(G)[2:3]
        G = '0'+ g
    if B > 15:
        B = ten_sixteen(B)[2:4]
    else:
        b = ten_sixteen(B)[2:3]
        B = '0' + b
    req = R+G+B
    return req

def creta_table():
    wb = Workbook()
    ws = wb.active
    im = Image.open("C:\Users\Administrator\Desktop\hsq.png")
    pix = im.load()
    width = im.size[0]
    height = im.size[1]
    print width,height
    for x in range(width):
        for y in range(height):
            print x, y
            colour = pix[x, y][0:3]
            #print colour
            R = colour[0]
            G = colour[1]
            B = colour[2]
            colour_code = Joining(R,G,B)
            row = int(width) - int(x)
            col = int(y) + 1
            print row,col
            ws.cell(row=col,column=row).fill=openpyxl.styles.PatternFill(fill_type='solid',fgColor=colour_code)

    wb.save('C:\Users\Administrator\Desktop\zm5.xlsx')

creta_table()

